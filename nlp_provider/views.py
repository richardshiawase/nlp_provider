import datetime
import json
import os
import pathlib
import pickle
import re
import shutil
from collections import defaultdict
import time
from functools import reduce
from io import BytesIO
from multiprocessing import Process, Pool
from time import sleep

import numpy as np
import requests
from django.core import serializers
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.forms import model_to_dict
from django.http import HttpResponse, HttpResponseRedirect, FileResponse
from django.shortcuts import render
import warnings

from fuzzywuzzy import fuzz
from openpyxl.styles import PatternFill

from classM.Asuransi import Asuransi
from classM.DFHandler import DFHandler
from classM.Dataset import Dataset
from classM.HospitalInsurance import HospitalInsurance
from classM.ItemMaster import ItemMaster
from classM.MasterData import MasterData
from classM.States import States
from model.models import ItemProvider, List_Processed_Provider, MatchProcess, MasterMatchProcess, GoldenRecordMatch
from classM.Pembersih import Pembersih
from classM.PerbandinganResult import PerbandinganResult
from classM.PredictionId import PredictionId
from classM.ColumnToRead import ColumnToRead

warnings.simplefilter(action='ignore', category=FutureWarning)
import pandas as pd
import pandas as pde
from requests import Response
from django.http import JsonResponse
import django

django.setup()
from sklearn import metrics
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics import precision_score, f1_score, accuracy_score

from model import models
from model.views import create_model

from model.models import Provider_Model, Provider
from tqdm import tqdm
from django.core.cache import cache

from celery import shared_task
import threading
# from .forms import UploadFileForm
# Create your views here.




ITERASI = 10
provider_dict_item = {}
master_item_list = []



server_prefix = "https://www.asateknologi.id"

class BackgroundTask(threading.Thread):
    def run(self):
        global master_data
        global tfidf_vec1
        global loaded_model1
        global state
        global asuransi
        global match_process
        global list_provider_model_object
        global golden_record_match
        global master_match_process

        golden_record_match = GoldenRecordMatch()
        golden_record_match.set_master_match_process()
        master_match_process = golden_record_match.get_master_match_process()

        match_process = MatchProcess()
        match_process.set_golden_record_instance(golden_record_match)
        match_process.start()
        match_process.set_list_provider()
        list_provider_model_object = match_process.get_list_provider()

        master_data = MasterData()
        print("{0} is loaded".format("Master Data"))

        state = States()
        print("{} is loaded".format("state"))

        asuransi = Asuransi()
        print("{} list is loaded".format("asuransi"))



t = BackgroundTask()
t.start()




def index(request):
    context = {"list_pembanding": []}
    return render(request, 'home.html', context)


def kompilasi(request):
    pembanding = models.Provider.objects.all()
    list_pembandinge = pembanding
    list_pembanding = []
    for pembanding in list_pembandinge:
        pembanding.file_location = pembanding.file_location.split("media")[1]
        list_pembanding.append(pembanding)

    return render(request, 'kompilasi.html')


def kompilasi_data(request):
    pembanding_all = models.Provider.objects.all()
    provider_list = []
    for pembanding in pembanding_all:
        pembanding.file_location = pembanding.file_location.split("media")[1]
        dfs = pd.read_excel("media/" + pembanding.file_location_result)
        for index, row in dfs.iterrows():
            alamat = row['Alamat']
            alamat_prediksi = row['Alamat Prediction']
            ri = row['ri']
            rj = row['rj']
            item_obj = ItemProvider(row['Provider Name'], row['Alamat'], row["Prediction"], row["Score"], 0, ri, rj)
            item_obj.set_nama_asuransi(pembanding.nama_asuransi)
            item_obj.set_selected(str(row['Compared']))
            item_obj.set_alamat_prediction(alamat_prediksi)

            provider_list.append(item_obj.__dict__)

    return JsonResponse(provider_list, safe=False)


def master_linked_load(request):
    # request.session['list_asuransi'] = asuransi
    if request.method == "GET":
        ls = asuransi.get_dict_item_asuransi().values()
        return JsonResponse(list(ls), safe=False)

    return JsonResponse({'message': 'error'})


def newe(request):
    list_provider_model_object.set_empty_provider_list()
    data_list = models.Provider.objects.raw(
        "select mm.id,mm.match_percentage,mm.id_model,mm.id_file_result,mm.status_finish,mm.created_at,mp.nama_asuransi,mp.file_location,mf.file_location_result from model_matchprocess mm inner join model_provider mp on mm.id_model = mp.id inner join model_fileresult mf on mm.id_file_result = mf.id  order by mm.created_at DESC")
    for data in data_list:
        provider = Provider()
        provider.set_nama_asuransi_model(data.nama_asuransi)
        provider.set_file_location(data.file_location)

        provider.set_id(data.id_model)
        provider.file_location_result = data.file_location_result
        provider.set_created_at(data.created_at)
        provider.set_status_matching(data.status_finish)
        list_item_provider = []
        list_item_provider_json = []
        dt = models.Provider.objects.raw("select * from model_itemprovider where id_model = %s",
                                         [provider.get_primary_key_provider()])
        for item in dt:
            item_provider = ItemProvider()
            item_provider.set_id(item.pk)
            item_provider.set_provider_name(item.nama_provider)
            item_provider.set_alamat_prediction(item.alamat_prediction)
            item_provider.set_alamat(item.alamat)
            item_provider.set_proba_score(item.proba_score)
            item_provider.set_total_score(item.total_score)
            item_provider.set_label_name(item.label_name)
            item_provider.set_ri(item.ri)
            item_provider.set_rj(item.rj)
            item_provider.set_id_asuransi(item.id_asuransi)
            item_provider.set_selected("-")
            item_provider.set_validity(item.validity)
            del item_provider._state

            # list_item_provider.append(item_provider)
            list_item_provider_json.append(item_provider.__dict__)
        provider.set_list_item_provider_json(list_item_provider_json)
        # provider.set_list_item_provider(list_item_provider)
        list_provider_model_object.add_provider(provider)

    if request.method == "GET":
        return JsonResponse(list_provider_model_object.get_provider_list_json(), safe=False)

    return JsonResponse({'message': 'error'})


def open_file_perbandingan(request):
    data = request.session['data_provider']
    id_provider = data['id']
    # provider = list_provider_model_object.get_a_provider_from_id(id_provider)
    return JsonResponse(data["list_item_provider_json"], safe=False)


def hos_ins_list_item(request):
    if 'hospital_linked_list' in request.session:
        hospital_linked_list = request.session['hospital_linked_list']
        return JsonResponse(list(hospital_linked_list.values()), safe=False)


def hos_ins_list_page(request):
    context = {"data": []}
    return render(request, 'matching/index_hospital_asuransi.html', context=context)


def hos_ins_list(request):
    asuransi_dict = asuransi.get_dict_item_asuransi()
    if request.method == "POST":
        data = json.loads(request.POST["data"])
        nama_asuransi = data["singkatan"]
        asu = asuransi_dict.get(data["singkatan"])

        request.session['nama_asuransi'] = nama_asuransi
        request.session['hospital_linked_list'] = asu["hospital_linked_list"]
        return JsonResponse({'data': asu["hospital_linked_list"], 'nama_asuransi': nama_asuransi})

def export_excel_linked_list(request):
    if request.method == "POST":
        # Get the raw JSON data from the request's body
        json_data = request.body.decode('utf-8')


        # Parse the JSON data into a Python dictionary
        data = json.loads(json_data)

        # Access the data and perform any necessary processing
        raw_data = data.get("data")
        print(raw_data)
        hospital_linked_dict_list = raw_data['hospital_linked_list']
        hospital_linked_list = hospital_linked_dict_list.values()

        result = None
        for item in hospital_linked_list:
            result = pd.concat([result, pd.DataFrame({
                'id_master': [item['hospital_id']],
                'nama_provider': [item['hospital_name']],
                'alamat': [item['hospital_address']]
            })], ignore_index=True)


        # Create a DataFrame
        # Export DataFrame to Excel
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        result.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        output.seek(0)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename='+raw_data['singkatan']+".xlsx"
        response.write(output.getvalue())

        return response



def update_master(request):
    if request.method == "POST":
        request.session["update_master"] = {"status":False,"object": {}}

        print("tes update master")
        data = json.loads(request.POST["processed_file"])
        row_index = data["row_index"]
        id_provider = data['id_provider']
        nama_provider = data["nama_provider"]
        alamat = data["alamat"]
        province = data["state"]
        city_choose = data["city"]
        telepon = data["telepon"]
        kategori = data["kategori"]
        latitude = data['latitude']
        longitude = data['longitude']

        state_id = "-"
        city_id = "-"
        kategori_id = 0

        kategori_dict = match_process.get_category_dict()

        for provinsi in state.get_item_state_dict().values():
            if province == provinsi.get_state_name():
                state_id = provinsi.get_state_id()

        for kota in state.get_item_city_list():
            if city_choose == kota.get_city_name():
                city_id = kota.get_city_id()

        for k, v in kategori_dict.items():
            print(kategori,k,v)
            if kategori == k:
                kategori_id = int(v)

        url = server_prefix+"/api/hospital/"+id_provider
        myobj = {'row_index':row_index,'id_provider':id_provider,'provider_name': nama_provider,'latitude':latitude,'longitude':longitude, 'address': alamat,'category_1':str(kategori_id),'tel_no':telepon,'state_id':str(state_id),'city_id':str(city_id)}
        x = requests.put(url, json=myobj)
        if x.status_code == 200:
            # request.session["update_master"] = {"status":True,"object":myobj}
            df_raw_master = master_data.get_dataframe()
            result = df_raw_master[df_raw_master['ProviderId'] == str(id_provider)].copy()
            # Update the selected columns in the filtered row

            result.loc[:, 'PROVIDER_NAME'] = nama_provider
            result.loc[:, 'ADDRESS'] = str(alamat)
            result.loc[:, 'TEL_NO'] = str(telepon)
            result.loc[:, 'Category_1'] = str(kategori_id)
            result.loc[:, 'stateId'] = str(state_id)
            result.loc[:, 'cityId'] = str(city_id)
            result.loc[:, 'latitude'] = str(latitude)
            result.loc[:, 'longitude'] = str(longitude)

            df_raw_master.update(result)
            print(df_raw_master.head())
            df_raw_master.to_excel("master_provider.xlsx", index=False)

            master_data.set_new_datafarame(df_raw_master)

            return JsonResponse({'data': 200})
        else:
            request.session["update_master"] = {"status":False,"object":[]}

            return JsonResponse({'data': 400})

        pass
    else:
        # print("tes update master")
        pass

    return JsonResponse({'data': 400})

def show_updated_master(request):
    print("Show updated master")
    message = {"status":False,"object":[]}
    if "update_master" in request.session:
        message = request.session.get("update_master")
        print(message)
        del request.session['update_master']
    # object = request.session["update_master"]["object"]
    return JsonResponse({'message': message})

def unlink_hos(request):
    if request.method == "POST":
        data = json.loads(request.POST["data"])
        id_hosins = data['id_hosins']
        id_asuransi = data['insurance_id']
        url = 'https://www.asateknologi.id/api/unlink-inshos'
        myobj = {'id_hosins': id_hosins,'id_asuransi':id_asuransi}
        x = requests.post(url, json=myobj)
        asuransi_dict = asuransi.get_dict_item_asuransi()
        asu = asuransi_dict.get(request.session['nama_asuransi'])

        if (x.status_code == 200):
            if 'hospital_linked_list' in request.session:
                hospital_linked_list = request.session['hospital_linked_list']
                del hospital_linked_list[str(id_hosins)]
                asu["hospital_linked_list"] = hospital_linked_list
                asu["linked_hospital_count"] = len(hospital_linked_list)
                request.session['hospital_linked_list'] = hospital_linked_list

        return JsonResponse({'data': x.status_code})


def perbandingan(request):
    if request.method == "POST":
        data = json.loads(request.POST["data"])
        request.session['data_provider'] = data
        # request.session['response'] = response
        # loop_delete(file_location)
        return JsonResponse(200, safe=False)


def perbandingan_page(request):
    data = request.session['data_provider']
    nama_asuransi = data['nama_asuransi']
    link_result = data["file_location_result"]

    context = {"list": [], "link_result": link_result, 'nama_asuransi': nama_asuransi}
    return render(request, 'matching/perbandingan_page_open_result.html', context=context)


def instant_search(request):
    return render(request, 'matching/instant_search.html')




def instant_search_process(request):
    if request.method == "POST":
        data = json.loads(request.POST["data"])

        y_preds, nil, score = match_process.calculate_score_instant(data)
        if nil < 0.5:
            y_preds = "Data belum dipelajari"

        return JsonResponse({'data': y_preds, 'nil': nil, 'score': score})


def perbandingan_upload_page(request):
    response = requests.get('https://asateknologi.id/api/insuranceall')
    response = response.json()
    context = {"list_insurance": response["val"]}
    return render(request, 'matching/perbandingan-upload.html', context=context)


def perbandingan_versus_page(request):
    response = requests.get('https://asateknologi.id/api/insuranceall')
    response = response.json()
    context = {"list_insurance": response["val"]}
    return render(request, 'matching/perbandingan-versus.html', context=context)


def tampungan(request):
    link_result = file_location
    if link_result is None:
        link_result = "-"

    context = {"provider_list": [], "link_result": link_result}
    return render(request, 'matching/perbandingan_basket.html', context=context)


def linked_master(request):
    return render(request, 'matching/linked_master.html')

    # return JsonResponse(asuransi.get_dict_item_asuransi(), safe=False)


def tampungan_rev(request):
    global provider_liste
    global file_location
    provider_liste = []

    dfs = None

    try:
        dfs = pd.read_excel("basket_provider.xlsx")
    except:
        print("dataframe not founde")

    provider_name_list = []
    provider_name_predict_list = []
    score_list = []
    # df_dataset = pd.read_excel("dataset_excel_copy.xlsx")

    provider_list = []
    if dfs is not None:
        for index, row in tqdm(dfs.iterrows(), total=dfs.shape[0]):

            provider_name_label = str(row['course_title']).strip().lower()
            alamat = str(row['alamat']).strip().lower()
            concat = provider_name_label + "#" + alamat
            concat = concat.replace('&', '').replace('.', '')
            sample1 = vectorize_text(concat, tfidf_vec1)
            y_preds = loaded_model1.predict(sample1)
            p = loaded_model1.predict_proba(sample1)
            ix = p.argmax(1).item()
            nil = (f'{p[0, ix]:.2}')
            # if(float(nil.strip("%")) < 1.0):
            # y_preds = "-"
            provider_name_list.append(provider_name_label)
            provider_name_predict_list.append(y_preds)
            score_list.append(nil)

            val = (df_non_duplicate['course_titles'].eq(provider_name_label))
            res = df_non_duplicate[val]
            provider_object = ItemProvider(provider_name_label, alamat, y_preds, nil, 0, 0, 0)

            if not res.empty:
                pred = str(y_preds).replace("[", "").replace("]", "").replace("'", "")
                val_master = (df_non_duplicate['subject'].eq(pred))
                res_master = df_non_duplicate[val_master]

                al = res_master["alamat"].head(1)
                try:
                    alamat_pred = al.values[0]
                except:
                    print("error")
            elif res.empty:
                alamat_pred = "-"

            provider_object.set_alamat_prediction(alamat_pred)

            provider_list.append(provider_object.__dict__)

    return JsonResponse(provider_list, safe=False)


def hapus_tampungan(request):
    dfs = pd.read_excel("basket_provider.xlsx")

    if request.method == "POST":
        nama = request.POST['nama_provider']
        # nama = json.load(request).get('dats')
        print(nama.upper())
        delete_row = dfs[dfs["course_title"] == nama.upper()].index
        dfs = dfs.drop(delete_row)
        # print(dfs)
        dfs.to_excel('basket_provider.xlsx', index=False)

    return HttpResponse("OK")


def upload_master(request):
    global provider_liste
    global file_location
    provider_liste = []
    response = requests.get('https://asateknologi.id/api/insuranceall')
    response = response.json()
    dfs = None

    if request.method == "POST":
        file_location = "media" + request.POST["file_location"]

    # elif request.method == "GET":
    # file_location="media/demo.xlsx"

    try:
        dfs = pd.read_excel(file_location)
    except:
        print("dataframe not found")
    provider_list = []
    # # # MASUKKAN DF KE LIST
    if dfs is not None:
        for index, row in dfs.iterrows():
            provider_name = row['Provider Name']
            alamat = row['Alamat']
            alamat_prediction = row['Alamat Prediction']
            y_preds = row["Prediction"]
            nil = row["Score"]
            ri = row["RI"]
            rj = row["RJ"]
            provider_object = ItemProvider(provider_name, alamat, y_preds, nil, 0, ri, rj)
            provider_object.set_alamat_prediction(alamat_prediction)
            provider_list.append(provider_object)
    page = request.GET.get('page', 1)
    paginator = Paginator(provider_list, 10)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    context = {"list_insurance": response.get("val"), "list": provider_list}

    return render(request, 'master/bulk_upload.html', context=context)


def list_master(request):
    return render(request, 'master/index.html')


def list_master_varian(request):
    return render(request, 'master/index_master_varian.html')


def list_master_sinkron(request):
    return render(request, 'master/sinkron.html')


def master_add(request):
    kategori_dict = match_process.get_category_dict()
    context = {"state_list": state.get_item_state_dict().values(), "city_list": state.get_item_city_list(),
               "kategori_dict": kategori_dict.keys()}
    return render(request, 'master/master_add.html', context=context)


def list_master_process(request):
    list_item_master = master_data.get_list_item_master_provider_json()
    return JsonResponse(list_item_master, safe=False)

def get_master_with_api():
    response = requests.get('https://asateknologi.id/api/daftar-rs-1234')
    provider_list = response.json().get("val")

    return provider_list


def sinkron_master_process(request):
    provider_list = get_master_with_api()

    master_data_list = []
    # master_data = MasterData()

    df = pd.DataFrame()

    for prov in provider_list:
        id = prov["id"]
        stateId = prov["stateId"]
        cityId = prov["CityId"]
        try:
            category_1 = int(prov["Category_1"])
        except:
            category_1 = 0
        category_2 = prov["Category_2"]
        telephone = prov["TEL_NO"]
        provider_name_master = prov["PROVIDER_NAME"]
        address = prov["ADDRESS"]
        latitude = prov['lat']
        longitude = prov['longitude']
        df = df.append(pd.Series(
            {'ProviderId': id, 'stateId': stateId, 'cityId': cityId, 'Category_1': category_1, 'Category_2': category_2,
             'PROVIDER_NAME': provider_name_master, 'ADDRESS': address, 'TEL_NO': telephone, 'latitude':latitude,'longitude':longitude},
            name=3))

    df.to_excel("master_provider.xlsx", index=False)

    return JsonResponse(master_data_list, safe=False)


def sinkron_master_process_not_request():
    print("Sinkron master proses")
    provider_list = get_master_with_api()

    master_data_list = []
    # master_data = MasterData()

    df = pd.DataFrame()

    for prov in provider_list:
        id = prov["id"]
        stateId = prov["stateId"]
        cityId = prov["CityId"]
        try:
            category_1 = int(prov["Category_1"])
        except:
            category_1 = 0
        category_2 = prov["Category_2"]
        telephone = prov["TEL_NO"]
        provider_name_master = prov["PROVIDER_NAME"]
        address = prov["ADDRESS"]

        df = df.append(pd.Series(
            {'ProviderId': id, 'stateId': stateId, 'cityId': cityId, 'Category_1': category_1, 'Category_2': category_2,
             'PROVIDER_NAME': provider_name_master, 'ADDRESS': address, 'TEL_NO': telephone},
            name=3))

    df.to_excel("master_provider.xlsx", index=False)

    return JsonResponse(master_data_list, safe=False)


def download_master(request):
    file_path = os.getcwd() + "\\master_provider.xlsx"
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(),
                                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=master_provider.xlsx'
            return response
    else:
        raise None
    return response


def download_master_varian(request):
    file_path = os.getcwd() + "\\master_varian_1.xlsx"
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(),
                                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=master_varian.xlsx'
            return response
    else:
        raise None
    return response


def sinkron_dataset_process(request):
    dataset_df = pd.DataFrame()

    find = False
    master_data_list = []
    dfs_varian = None
    # try:
    #     dataset_df = cache.get('dataset')
    #     if dataset_df is None:
    #         dataset_df = pd.read_excel("dataset_excel_copy.xlsx")
    #         cache.set('dataset', dataset_df)
    #     dfs_varian = dataset_df.groupby('subject')
    # except:
    #     print("dataframe not found")
    for item_master in master_data.get_dict_item_master_provider().values():
        provider_name_master = item_master.get_nama_master()
        alamat = item_master.get_alamat_master()
        varian_list = []

        try:
            dfe = dfs_varian.get_group(provider_name_master)
            for index_varian, row_varian in dfe.iterrows():
                varian_list.append(row_varian['course_title'])
                pass

        except:
            row = pd.Series(
                {'course_title': provider_name_master, 'alamat': alamat,
                 'subject': provider_name_master},
                name=3)
            dataset_df = dataset_df.append(row)
            dataset_df.reset_index(drop=True, inplace=True)

            continue

    dataset_df.to_excel("dataset_excel_copy.xlsx", index=False)
    return HttpResponse("Tes")


def master_varian_process(request):
    dff = pd.DataFrame()
    # existing_series = pd.Series()
    dataset = match_process.get_dataset()
    master_data_list = []

    # SET VARIAN FROM SUBJECT COLUMN OF DATASET
    dataset_grouped_by_subject = dataset.get_bulk_dataset().groupby('subject')

    for item_master in tqdm(master_data.get_list_item_master_provider(),total=len(master_data.get_list_item_master_provider())):
        varian_list = []
        try:
            dfe = dataset_grouped_by_subject.get_group(item_master.get_nama_master())
            for index_varian, row_varian in dfe.iterrows():
                nama_varian = row_varian['course_title']
                alamat = row_varian['alamat']
                nama_varian_alamat = nama_varian+"#"+alamat
                varian_list.append(nama_varian_alamat)
        except Exception as e:
            # print(e)
            continue

        item_master.set_varian(varian_list)

        try:
            new_data = {'ProviderId': item_master.get_id_master(), 'ProviderType': "Master",
                 'stateId': item_master.get_state_id_master(), 'cityId': item_master.get_city_id_master(),
                 'Category_1': item_master.get_category_1_master(),
                 'Category_2': item_master.get_category_2_master(),
                 'PROVIDER_NAME': item_master.get_nama_master(), 'ADDRESS': item_master.get_alamat_master(),
                 'TEL_NO': item_master.get_telepon_master()}
            dff = dff.append(new_data, ignore_index=True)
        except Exception as e:
            print(e)

        for varian in item_master.get_varian():
            try:
                nama = varian.split("#")[0]
                alamat = varian.split("#")[1]
                varian_data = {'ProviderId': item_master.get_id_master(), 'ProviderType': "Varian",
                     'stateId': item_master.get_state_id_master(), 'cityId': item_master.get_city_id_master(),
                     'Category_1': item_master.get_category_1_master(),
                     'Category_2': item_master.get_category_2_master(),
                     'PROVIDER_NAME': nama, 'ADDRESS': alamat, 'TEL_NO': "-"}
                dff = dff.append(varian_data, ignore_index=True)
            except Exception as e:
                print(e)



    dff.to_excel("master_varian_1.xlsx", index=False)

    return JsonResponse(master_data_list, safe=False)


def master_varian_list_read(request):
    dff = pd.DataFrame()

    find = False
    master_data_list = []
    dfs = None
    try:
        dfs = pd.read_excel("master_varian_1.xlsx")
    except:
        print("dataframe not found")

    for index, row in dfs.iterrows():
        id = row['ProviderId']
        stateId = row['stateId']
        cityId = row['cityId']
        category_1 = row['Category_1']
        category_2 = row['Category_2']
        provider_name_master = row['PROVIDER_NAME']
        address = row['ADDRESS']
        tel_no = row['TEL_NO']
        master_data = MasterData(id, provider_name_master, address, category_1, category_2, tel_no, stateId, cityId)
        varian_list = []

        try:
            dfe = dfs_varian.get_group(provider_name_master)
            for index_varian, row_varian in dfe.iterrows():
                varian_list.append(row_varian['course_title'])
                pass

        except:
            continue

        master_data.set_varian(varian_list)

        dff = dff.append(pd.Series(
            {'ProviderId': id, 'ProviderType': "Master", 'stateId': stateId, 'cityId': cityId, 'Category_1': category_1,
             'Category_2': category_2,
             'PROVIDER_NAME': provider_name_master, 'ADDRESS': address, 'TEL_NO': tel_no},
            name=3))

        for varian in master_data.get_varian():
            dff = dff.append(pd.Series(
                {'ProviderId': id, 'ProviderType': "Varian", 'stateId': stateId, 'cityId': cityId,
                 'Category_1': category_1, 'Category_2': category_2,
                 'PROVIDER_NAME': varian, 'ADDRESS': "-", 'TEL_NO': "-"},
                name=3))

        master_data_list.append(master_data.__dict__)

    #
    dff.to_excel("master_varian_1.xlsx", index=False)

    return JsonResponse(master_data_list, safe=False)


def temporer_store(request):
    if request.method == "POST":

        id = request.POST['id']
        if id is not None and id != '':
            item = ItemProvider.objects.get(pk=id)

            if id in provider_dict_item:
                del provider_dict_item[id]
            else:
                provider_dict_item[id] = item

    context = {"provider_name": provider_dict_item, "link_result": "-"}

    # return HttpResponse(context)
    return render(request, 'matching/temporer.html', context=context)

def temporer_store_master(request):
    if request.method == "POST":
        master_item_list.clear()
        master = json.loads(request.POST['processed_file'])
        row_index = master['row_index']
        id      = master['provider_id']
        state_id = master['stateId']
        city_id = master['cityId']
        category_1 = master['category_1']
        category_2 = master['category_2']
        provider_name = master['provider_name']
        address = master['address']
        tel_no = master['tel_no']
        latitude = master['lat']
        longitude = master['longitude']
        item_master = ItemMaster(id,
                                 state_id,
                                 city_id,
                                 category_1,
                                 category_2,
                                 provider_name,
                                 address,
                                 tel_no)
        item_master.set_master_latitude(latitude)
        item_master.set_master_longitude(longitude)
        print(item_master)

        item_master.set_datatable_row_index(row_index)
        master_item_list.append(item_master)
    if len(master_item_list) > 0:
        return JsonResponse({'data':200,'link':'-'})
    return JsonResponse({'data':404,'link':'-'})


def temporer_delete_master(request):
    if request.method == "POST":
        master_item_list.clear()
        master = json.loads(request.POST['processed_file'])
        row_index = master['row_index']

        id      = master['provider_id']
        state_id = master['stateId']
        city_id = master['cityId']
        category_1 = master['category_1']
        category_2 = master['category_2']
        provider_name = master['provider_name']
        address = master['address']
        tel_no = master['tel_no']

        item_master = ItemMaster(id,
                                 state_id,
                                 city_id,
                                 category_1,
                                 category_2,
                                 provider_name,
                                 address,
                                 tel_no)

        item_master.set_datatable_row_index(row_index)
        master_item_list.append(item_master)

        url = server_prefix + "/api/hospital/" + id

        x = requests.delete(url)

        if x:
            return JsonResponse({'data':200,'link':'-'})
    return JsonResponse({'data':404,'link':'-'})

def edit_master(request):
    if request.method == "POST":
        print("Googl")
    else:
        item_master = master_item_list[0]
    kategori_dict = match_process.get_category_dict()
    state_name = "-"
    city_name = "-"
    kategori = "-"
    for provinsi in state.get_item_state_dict().values():
        if int(item_master.get_state_id_master()) == int(provinsi.id):
            state_name = provinsi.state_name

    for kota in state.get_item_city_list():
        if int(item_master.get_city_id_master()) == int(kota.get_city_id()):
            city_name = kota.get_city_name()

    for k,v in kategori_dict.items():
        if int(v) == int(item_master.get_category_1_master()):
            kategori = k

    context = {"item_master": item_master, "link_result": "-","state_list": state.get_item_state_dict().values(), "city_list": state.get_item_city_list(),
               "kategori_dict": kategori_dict.keys(),"state_name":state_name,"city_name":city_name,"kategori":kategori}
    #
    # return "ooops"
    return render(request, 'matching/edit_master.html', context=context)

def read_link_result_and_delete_provider_name2(nama_provider, link_result):
    global dfs

    val = (dfs['Nama'].str.lower().eq(nama_provider.lower()))
    rese = dfs[val]

    # if not rese.empty:
    #     print(rese.index.item())
    # print(dfs)
    global deo
    global deoq
    if not rese.empty:
        #

        try:
            deo = dfs.drop(rese.index.item(), inplace=True)

            val = (dw['Nama'].str.lower().eq(nama_provider.lower()))
            reseq = dw[val]
            if not reseq.empty:
                deoq = dw.drop(reseq.index.item(), inplace=True)
                # deoq = dw
                # if (nama_provider == "klinik takenoko sudirman"):
                #     vae = deoq['Nama Provider'].str.strip().str.lower().eq("klinik takenoko sudirman")
                #     print(deoq[vae])
        except Exception as e:
            for x in rese.index.tolist():
                deo = dfs.drop(x, inplace=True)

            pass


def read_link_result_and_delete_provider_name(nama_provider):
    dfs = pd.read_excel(link_result)
    val = (dfs['Provider Name'].eq(nama_provider.upper()))
    rese = dfs[val]
    print("hapus1 " + nama_provider, rese)

    if not rese.empty:
        print("hapus " + nama_provider, rese)
        deo = dfs.drop(rese.index.item())
        deo.to_excel(link_result, sheet_name='Sheet1', index=False)
        dat = Provider.objects.filter(file_location_result__contains=link_result.split("/")[1]).values()
        dw = pd.read_excel(dat[0]["file_location"])
        val = (dw['Nama Provider'].eq(nama_provider.upper()))
        reseq = dw[val]
        if not reseq.empty:
            deoq = dw.drop(reseq.index.item())
            deoq.to_excel(dat[0]["file_location"], sheet_name='Sheet1', index=False)


def loop_delete(link_result):
    print("loop data2 ", link_result)
    global dfs
    global deo
    global deoq
    global dw
    deo = None
    deoq = None

    file_master = "Master_Add.xlsx"
    df_handler.convert_to_dataframe_from_excel(file_master)
    df = df_handler.get_data_frame()

    dat = Provider.objects.filter(file_location_result__contains=link_result.split("/")[1]).values()
    file_location = dat[0]["file_location"]

    df_handler.convert_to_dataframe_from_excel(file_location)
    dw = df_handler.get_data_frame()

    df_handler.convert_to_dataframe_from_excel(link_result)
    dfs = df_handler.get_data_frame()

    for index, row in tqdm(df.iterrows(), total=df.shape[0]):
        nama_provider = row['provider_name']
        read_link_result_and_delete_provider_name2(nama_provider, link_result)

    dfs.to_excel(link_result, sheet_name='Sheet1', index=False)




class BackgroundTaskDataset(threading.Thread):
    def run(self):
        item_master = self.item_master
        cache.delete('dataset')
        match_process.set_dataset()
        dataset = match_process.get_dataset()
        df = dataset.get_bulk_dataset()

        print(df['updated'])
        for x in range(ITERASI):
            row = pd.Series(
                {'course_title': item_master.get_nama_master(), 'alamat': item_master.get_alamat_master(),
                 'updated': int(0),
                 'subject': item_master.get_nama_master()}, name=3)
            df = df.append(row, ignore_index=True)

        pembersih = Pembersih(df)
        df = pembersih._return_df()
        print(df['updated'])
        df.to_excel("dataset_excel_copy.xlsx", index=False)
        match_process.set_dataset()

    def set_item_master(self,item_master):
        self.item_master



def add_master_by_dashboard(request):
    if request.method == "POST":
        nama_provider = request.POST["nama_provider"]
        alamat_provider = request.POST["alamat_provider"]
        provinsi_provider = request.POST["provinsi_provider"]
        city_provider = request.POST["city_provider"]
        telepon_provider = request.POST["telepon_provider"]
        kategori_provider = request.POST["kategori_provider"]
        latitude_provider = request.POST["latitude_provider"]
        longitude_provider = request.POST["longitude_provider"]
        cat_dict = match_process.get_category_dict()
        kategori_provider = cat_dict.get(kategori_provider)
        provinsi_provider = state.get_item_state_dict().get(provinsi_provider)
        city = state.get_city()
        city_provider = city.get_item_city_only_dict().get(city_provider)

        item_master = ItemMaster("-",provinsi_provider.get_state_id(),city_provider.get_city_id(),kategori_provider,"-",nama_provider,alamat_provider,telepon_provider)
        url = 'https://www.asateknologi.id/api/master'
        myobj = {'stateId': provinsi_provider.get_state_id(),
                 'cityId': city_provider.get_city_id(),
                 'category1': kategori_provider,
                 'provider_name': nama_provider,
                 'address': alamat_provider,
                 'tel': telepon_provider,
                 'latitude': latitude_provider,
                 'longitude': longitude_provider
                 }
        try:
            pass
            x = requests.post(url, json=myobj)
            token = "eyJ0eXBlIjoiSldUIiwiYWxnIjoiSFMyNTYifQ.eyJzdWIiOiJ1c2VyZm9ycHJvdmlkZXIiLCJpYXQiOjE2ODMyNzExNjYsIm5hbWUiOiJ1c2VyZm9ycHJvdmlkZXIifQ.l65gkzEqH-uuN9b84ZU4aADwM2Rb3nZRgsmmAqwTQsc"
            header = {"Content-Type": "application/json", "Authorization": f"Bearer {token}"}
            url_sinkron_sinta = "http://192.168.80.210/be/api/dashboard/syncronize"
            d = requests.get(url_sinkron_sinta,headers=header)

        except Exception as e:
            print(e)

        try:

            # # # CLEAR ALL DATASET CACHE
            # # # READ DATASET FRESHLY FROM EXCEL
            # # # ADD TO DATASET
            # # # SAVE TO DATASET AND CREATE CACHE
            dt = BackgroundTaskDataset()

            dt.set_item_master(item_master)
            dt.start()

        except Exception as e:
            print(e)

    return HttpResponse(200)


def add_master_store(request):
    if request.method == "POST":
        df = pd.read_excel("Master_Add.xlsx")
        post_ide = request.POST["post_idew"]
        nama_provider = post_ide.split("#")[0]
        alamat = post_ide.split("#")[1]
        link_result = request.POST["link_result"]
        val = (df['provider_name'].str.lower().eq(nama_provider.lower()))
        res = df[val]
        # # # kalau kosong alias belum ada nama provider di dalam file master add, maka proses
        if res.empty:
            row = pd.Series({'provider_name': nama_provider, 'alamat': alamat})
            df = df.append(row, ignore_index=True)
            df.to_excel("Master_Add.xlsx", index=False)

        read_link_result_and_delete_provider_name(nama_provider)

        # # # tambahin ke dataset

    else:
        return HttpResponse("OK")

    # return HttpResponse(context)
    return HttpResponse("OK")


def update_temporer_store(request):
    global name
    global link_result
    if request.method == "POST":
        post_ide = request.POST["post_idew"]
        link_result = request.POST["link_result"]
        name = post_ide
        context = {"provider_name": post_ide, "link_result": link_result}
    else:
        if name in provider_liste:
            provider_liste.remove(name)
        context = {"provider_name": provider_liste, "link_result": link_result}
    return render(request, 'matching/temporer.html', context=context)


def add_to_dataset(request):
    if request.method == "POST":
        # OPEN DATASET FILE
        dataset = match_process.get_dataset()
        df = dataset.get_bulk_dataset()

        df_basket = pd.read_excel("basket_provider.xlsx")

        # SEARCH PROVIDER IN DATASET
        for label_name, key_provider in list(
                zip(request.POST.getlist('nama_label'), request.POST.getlist('value_provider'))):
            label_name = label_name.split("#")[0]

            # key provider is key of item provider that being fixed
            # use key provider to search item provider that being fixed
            item_provider = ItemProvider.objects.get(pk=key_provider)

            ctr = 0
            for index, row in df.iterrows():
                label = row["subject"]
                if label == label_name:
                    if row['updated'] == str(0):
                        if ctr > 0 :
                            df.at[index, 'course_title'] = item_provider.get_nama_provider()
                            df.at[index, 'alamat'] = item_provider.get_alamat()
                            df.at[index, 'updated'] = str(1)
                            df.at[index, 'subject'] = label_name
                            break
                    ctr += 1
            try:
                rowe = pd.Series(
                    {'course_title': item_provider.get_nama_provider(), 'alamat': item_provider.get_alamat(),'subject': label_name}, name=3)
                df_basket = df_basket.append(rowe, ignore_index=True)
            except:
                break

        df_basket.to_excel("basket_provider.xlsx", index=False)
        df.to_excel("dataset_excel_copy.xlsx", index=False)

        context = {"list_pembanding": []}

        return render(request, 'home.html', context)

    return HttpResponse("Marco Polo")


def process_temporer_store(request):
    # dfs = cache.get('dataset')
    # if dfs is None:
    #     dfs = pd.read_excel("dataset_excel_copy.xlsx")
    #     cache.set('dataset', dfs)
    #
    # dfz = dfs.dropna(subset="alamat")
    # dfa = dfz.drop_duplicates(subset='subject')
    label_list = []
    dataset = match_process.get_dataset()
    dfa = dataset.get_dataframe_after_cleaned_no_duplicate()
    for index, row in dfa.iterrows():
        alamat = str(row['alamat'])
        label = row["subject"]
        if label + "#" + alamat not in label_list:
            label_list.append(label + "#" + alamat)
    context = {"label_list": label_list, "list": provider_dict_item, "link_result": "-"}
    # return HttpResponse("Process Temporer")
    return render(request, 'matching/proses_temporer.html', context=context)


def get_label(request):
    dfs = cache.get('dataset')
    if dfs is None:
        dfs = pd.read_excel("dataset_excel_copy.xlsx")
        cache.set('dataset', dfs)
    dfs = dfs.sort_values(by=['subject'], ascending=True)
    dfz = dfs.dropna(subset="alamat")
    dfa = dfz.drop_duplicates(subset='subject')
    label_list = []
    print(dfa.size, dfs.size)
    for index, row in dfa.iterrows():
        provider_name = row['course_title']
        alamat = str(row['alamat'])
        label = row["subject"]
        if label + "#" + alamat not in label_list:
            label_list.append(label + "#" + alamat)

    context = {"label_list": label_list}
    return JsonResponse(context, safe=False)


def check_header(df):
    header_list = ['Provinsi', 'Kota', 'Nama Provider', 'Alamat']
    df_header_list = list(df.columns.values)
    if df_header_list == header_list:
        return True
    return False


def vectorize_text(text, tfidf_vec):
    # text = "Klinik Ananda"
    my_vec = tfidf_vec.transform([text])
    return my_vec.toarray()


def cacah_dataframe(df):
    split_row_each = 800
    start_index = 0
    iteration_count = int(df.shape[0] / split_row_each)
    sisa = df.shape[0] % split_row_each
    sisa_row = iteration_count * split_row_each + sisa
    df_list = []
    for x in range(iteration_count):
        end_index = start_index + split_row_each
        df_new = df.iloc[start_index:end_index]
        start_index = end_index
        # df_list.append([df_new,lr])
        df_list.append(df_new)
    aw = lambda x, y: y if x > 0 else 0
    df_last = df.iloc[start_index:aw(sisa, sisa_row)]
    df_list.append(df_last)

    return df_list


def is_file_with_this_insurance_exists(nama_asuransi):
    mydata = Provider.objects.filter(nama_asuransi__contains=nama_asuransi).order_by('created_at').values()
    return mydata


def update_perbandingan_excel():
    pass


def perbandingan_result(request):
    global uploaded_file
    global contexte
    global perbandingan_model
    # sinkron_master_process_not_request()

    if request.method == 'POST':
        # # # REQUEST DARI PROSES FILE
        if not bool(request.FILES.get('perbandinganModel', False)):
            pembanding_model_return = json.loads(request.POST['processed_file'])
            nama_asuransi = pembanding_model_return["nama_asuransi"]
            provider = Provider.get_model_from_filter(nama_asuransi)

        # # # REQUEST DARI UPLOAD FILE
        else:
            # # init file storage object
            file_storage = FileSystemStorage()

            # # init Perbandingan object
            provider = Provider()

            # # get nama asuransi and file request
            data_asuransi = request.POST['insurance_option']
            file = request.FILES['perbandinganModel']

            nama_asuransi = str(data_asuransi).split("#")[0]
            id_asuransi = str(data_asuransi).split("#")[1]
            # save the file to /media/
            c = file_storage.save(file.name, file)

            # get file url
            file_url = file_storage.path(c)

            # set file location and nama_asuransi to Perbandingan object
            provider.set_file_location(file_url)
            provider.set_nama_asuransi_model(nama_asuransi)
            provider.set_id_asuransi_model(id_asuransi)
            provider.link_to_item_list()

        list_provider_model_object.add_provider(provider)
        start_time = time.time()
        match_process.set_master_data(master_data)
        match_process.process_matching()
        match_process.create_file_result()

        master_match_process.set_master_data(master_data)
        master_match_process.set_file_result_match_processed(match_process.get_file_result())
        master_match_process.process_master_matching()
        master_match_process.save_matching_information()

        golden_record_match.set_final_result(master_match_process.get_file_final_result_master_match())
        golden_record_match.set_file_result(master_match_process.get_file_result_match_processed())
        golden_record_match.process_golden_record()
        # master_match_process.delete_provider_item_hospital_insurances_with_id_insurances()
        # master_match_process.insert_into_end_point_andika_assistant_item_provider()
        print("--- %s seconds ---" % (time.time() - start_time))

        list_item_provider_json = []
        list_item_provider = []

        dt = models.Provider.objects.raw("select * from model_itemprovider where id_model = %s",
                                         [provider.get_primary_key_provider()])
        for item in dt:
            item_provider = ItemProvider()
            item_provider.set_id(item.pk)
            item_provider.set_provider_name(item.nama_provider)
            item_provider.set_alamat_prediction(item.alamat_prediction)
            item_provider.set_alamat(item.alamat)
            item_provider.set_proba_score(item.proba_score)
            item_provider.set_total_score(item.total_score)
            item_provider.set_label_name(item.label_name)
            item_provider.set_ri(item.ri)
            item_provider.set_rj(item.rj)
            item_provider.set_id_asuransi(item.id_asuransi)
            item_provider.set_selected("-")
            del item_provider._state

            list_item_provider.append(item_provider)
            list_item_provider_json.append(item_provider.__dict__)
        provider.set_list_item_provider_json(list_item_provider_json)
        provider.set_list_item_provider(list_item_provider)

    return JsonResponse(provider.get_list_item_provider_json(), safe=False)
def perbandingan_result_versus(request):
    global uploaded_file
    global contexte
    global perbandingan_model
    sinkron_master_process_not_request()
    master_data = MasterData()
    master_df = []
    print(os.getcwd())
    if request.method == 'POST':
        data_provider = []
        # # # REQUEST DARI PROSES FILE
        if not bool(request.FILES.get('perbandinganModel1', False)):
            pembanding_model_return = json.loads(request.POST['processed_file'])
            nama_asuransi = pembanding_model_return["nama_asuransi"]
            provider = Provider.get_model_from_filter(nama_asuransi)

        # # # REQUEST DARI UPLOAD FILE
        else:
            # # init file storage object
            file_storage = FileSystemStorage()

            # # init Perbandingan object
            provider1 = Provider()
            provider2 = Provider()

            # # get nama asuransi and file request
            data_asuransi1 = request.POST['insurance_option1']
            data_asuransi2 = request.POST['insurance_option2']



            file1 = request.FILES['perbandinganModel1']
            file2 = request.FILES['perbandinganModel2']




            nama_asuransi1 = str(data_asuransi1).split("#")[0]
            nama_asuransi2 = str(data_asuransi2).split("#")[0]

            id_asuransi1 = str(data_asuransi1).split("#")[1]
            id_asuransi2 = str(data_asuransi2).split("#")[1]




            # save the file to /media/
            c1 = file_storage.save(file1.name, file1)
            c2 = file_storage.save(file2.name, file2)


            # get file url
            file_url1 = file_storage.path(c1)
            file_url2 = file_storage.path(c2)


            # set file location and nama_asuransi to Perbandingan object
            provider1.set_file_location(file_url1)
            provider2.set_file_location(file_url2)
            print(nama_asuransi1,nama_asuransi2)
            provider1.set_nama_asuransi_model(nama_asuransi1)
            provider2.set_nama_asuransi_model(nama_asuransi2)

            provider1.set_id_asuransi_model(id_asuransi1)
            provider2.set_id_asuransi_model(id_asuransi2)

            provider1.link_to_item_list()
            provider2.link_to_item_list()

            data_provider.append(provider1)
            data_provider.append(provider2)


        for provider in data_provider:
            start_time = time.time()
            match_process.set_master_data(master_data)
            match_process.process_matching_versus(provider)
            match_process.create_file_result()
            master_match_process.set_master_data(master_data)
            master_match_process.set_file_result_match_processed(match_process.get_file_result())
            master_match_process.process_master_matching()


            master_df.append(master_match_process.get_final_result_dataframe())

            master_match_process.save_matching_information()

            golden_record_match.set_final_result(master_match_process.get_file_final_result_master_match())
            golden_record_match.set_file_result(master_match_process.get_file_result_match_processed())
            golden_record_match.process_golden_record()
            print("--- %s seconds ---" % (time.time() - start_time))

            list_item_provider_json = []
            list_item_provider = []

            dt = models.Provider.objects.raw("select * from model_itemprovider where id_model = %s",
                                             [provider.get_primary_key_provider()])
            for item in dt:
                item_provider = ItemProvider()
                item_provider.set_id(item.pk)
                item_provider.set_provider_name(item.nama_provider)
                item_provider.set_alamat_prediction(item.alamat_prediction)
                item_provider.set_alamat(item.alamat)
                item_provider.set_proba_score(item.proba_score)
                item_provider.set_total_score(item.total_score)
                item_provider.set_label_name(item.label_name)
                item_provider.set_ri(item.ri)
                item_provider.set_rj(item.rj)
                item_provider.set_id_asuransi(item.id_asuransi)
                item_provider.set_selected("-")
                del item_provider._state

                list_item_provider.append(item_provider)
                list_item_provider_json.append(item_provider.__dict__)
            provider.set_list_item_provider_json(list_item_provider_json)
            provider.set_list_item_provider(list_item_provider)

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    master_df[0]['Compared'] = 'False'
    master_df[1]['Compared'] = 'False'

    print("Comparing")
    for index,row in master_df[0].iterrows():
        id_master1 = int(row['IdMaster'])
        for index2,row2 in master_df[1].iterrows():
            id_master2 = int(row2['IdMaster'])

            if(id_master1 == id_master2):
                master_df[0].at[index, 'Compared'] = 'True'
                master_df[1].at[index2, 'Compared'] = 'True'
                break

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    output_1 = 'media\\file1_'+timestamp+'.xlsx'
    output_2 = 'media\\file2_'+timestamp+'.xlsx'
    master_df[0].to_excel(output_1, index=False, header=True)
    master_df[1].to_excel(output_2, index=False, header=True)


    # Create an Excel writer using pandas and openpyxl
    # excel_file_path = 'dataframe_data_all_red_rows.xlsx'
    # with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    #     master_df[0].to_excel(writer, sheet_name='Sheet1', index=False, header=True)
    #     worksheet = writer.sheets['Sheet1']
    #     for row_idx in range(2, len(master_df[0]) + 2):  # Start from row 2 (header is row 1)
    #         for col_idx in range(1, len(master_df[0].columns) + 1):
    #             worksheet.cell(row=row_idx, column=col_idx).fill = red_fill



    # for index,row in master_df[0].iterrows():
    #     id_master1 = row['IdMaster']
    #     for index,row in master_df[1].iterrows():
    #         id_master2 = row['IdMaster']


    # return HttpResponse(data_provider)
    request.session['output1'] = os.getcwd()+'\\'+output_1
    request.session['output2'] = os.getcwd()+'\\'+output_2
    return JsonResponse({'data':200,'link1':output_1,'link2':output_2})


def download_file(request):
    file1_path = request.session['output1']  # Replace with the actual file path
    file2_path = request.session['output2']  # Replace with the actual file path

    if os.path.exists(file1_path) and os.path.exists(file2_path):
        # Open both files in binary mode
        file1 = open(file1_path, 'rb')
        file2 = open(file2_path, 'rb')

        # Create a zip file containing both files
        response = HttpResponse(content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="files.zip"'

        import zipfile
        with zipfile.ZipFile(response, 'w') as zipf:
            zipf.writestr('file1.xlsx', file1.read())
            zipf.writestr('file2.xlsx', file2.read())

        # Close the files
        file1.close()
        file2.close()

        return response
    else:
        return HttpResponse("File not found", status=404)